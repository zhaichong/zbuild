# -*- coding: utf-8 -*-
"""Utility for creating order output directories and Excel test submission forms based on the official template.
"""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

logger = logging.getLogger(__name__)


def find_template_file(order_dir_base: Union[str, Optional[Path]] = None) -> Optional[Path]:
    """Find the official 提测单 .xlsx template file."""
    # 1. Check local references directory inside zbuild
    pkg_dir = Path(__file__).resolve().parents[2]
    bundled_tmpl = pkg_dir / "references" / "template_test_order.xlsx"
    if bundled_tmpl.exists():
        return bundled_tmpl

    # Also check relative to scripts
    scripts_dir = Path(__file__).resolve().parents[1]
    bundled_tmpl2 = scripts_dir / "references" / "template_test_order.xlsx"
    if bundled_tmpl2.exists():
        return bundled_tmpl2

    # Check via core.constants if available
    try:
        from core.constants import APP_DIR, DATA_DIR
        for cand in [
            APP_DIR / "references" / "template_test_order.xlsx",
            DATA_DIR / "references" / "template_test_order.xlsx",
            DATA_DIR / "template_test_order.xlsx",
        ]:
            if cand.exists():
                return cand
    except Exception:
        pass

    # 2. Check canonical locations in order_dir_base if provided
    if order_dir_base:
        base = Path(order_dir_base)
        if base.exists():
            for cand in [
                base.parent / "升级说明+提测单" / "template_test_order.xlsx",
                base.parent / "升级说明+提测单" / "医院提测单.xlsx",
                base / "升级说明+提测单" / "template_test_order.xlsx",
                base / "template_test_order.xlsx",
                base / "2026-1319 -深圳市新华医院" / "2026-1319 -深圳市新华医院医院提测单.xlsx",
            ]:
                if cand.exists():
                    return cand

    return None


def find_docx_template_file(order_dir_base: Union[str, Optional[Path]] = None) -> Optional[Path]:
    """Find the official 全部升级说明.docx template file."""
    # 1. Check bundled references inside zbuild
    pkg_dir = Path(__file__).resolve().parents[2]
    bundled_docx = pkg_dir / "references" / "全部升级说明.docx"
    if bundled_docx.exists():
        return bundled_docx

    scripts_dir = Path(__file__).resolve().parents[1]
    bundled_docx2 = scripts_dir / "references" / "全部升级说明.docx"
    if bundled_docx2.exists():
        return bundled_docx2

    # Check via core.constants if available
    try:
        from core.constants import APP_DIR, DATA_DIR, UPGRADE_DOC_NAME
        for cand in [
            APP_DIR / "references" / UPGRADE_DOC_NAME,
            APP_DIR / UPGRADE_DOC_NAME,
            DATA_DIR / "references" / UPGRADE_DOC_NAME,
            DATA_DIR / UPGRADE_DOC_NAME,
        ]:
            if cand.exists():
                return cand
    except Exception:
        pass

    # 2. Check exact user specified path
    exact_p = Path(r"D:\yh\特殊订单\升级说明+提测单\全部升级说明.docx")
    if exact_p.exists():
        return exact_p

    # 3. Check in order_dir_base canonical surrounding paths
    if order_dir_base:
        base = Path(order_dir_base)
        for cand in [
            base.parent / "升级说明+提测单" / "全部升级说明.docx",
            base / "升级说明+提测单" / "全部升级说明.docx",
            base / "全部升级说明.docx",
        ]:
            if cand.exists():
                return cand

    return None


def _format_change_notes(order_notes:Optional[str], projects: List[Dict[str, Any]]) -> str:
    """Format change notes for cell B12 of the Excel test submission form."""
    if order_notes and str(order_notes).strip():
        import re
        raw_lines = [line.strip() for line in str(order_notes).strip().splitlines() if line.strip()]
        formatted = []
        for idx, line in enumerate(raw_lines, start=1):
            if re.match(r"^\d+[、.．\s]", line):
                formatted.append(line)
            else:
                formatted.append(f"{idx}、{line}")
        return "\n".join(formatted)

    lines = []
    if projects:
        for idx, p in enumerate(projects, start=1):
            p_name = p.get("name") or p.get("projectName") or ""
            p_branch = p.get("branch") or p.get("currentBranch") or ""
            if p_branch:
                lines.append(f"{idx}、{p_name} ({p_branch})")
            else:
                lines.append(f"{idx}、{p_name}")
    else:
        lines.append("1、特殊订单需求更新")
    return "\n".join(lines)


def create_order_directory(
    order_dir_base: Union[str, Path],
    order_no: str,
    hospital_name: str,
    projects:Optional[List[Dict[str, Any]]] = None,
    order_notes:Optional[str] = None,
) -> Dict[str, Any]:
    """Create the order directory, test submission Excel file (.xlsx) and upgrade docx file (.docx).

    Naming rule:
      Folder: {order_dir_base}/{order_no}-{hospital_name}
      Excel:  {folder_name}医院提测单.xlsx
      Docx:   全部升级说明.docx
    """
    if not order_dir_base:
        return {"success": False, "message": "未配置提测目录根路径"}
    if not order_no or not hospital_name:
        return {"success": False, "message": "订单号和医院名称不能为空"}

    base_path = Path(order_dir_base)
    try:
        base_path.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        return {"success": False, "message": f"创建基础目录失败: {e}"}

    clean_order = str(order_no).strip()
    clean_hosp = str(hospital_name).strip()

    folder_name = f"{clean_order}-{clean_hosp}"
    target_dir = base_path / folder_name

    try:
        target_dir.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        return {"success": False, "message": f"创建订单目录失败 [{target_dir}]: {e}"}

    excel_name = f"{folder_name}医院提测单.xlsx"
    excel_path = target_dir / excel_name

    excel_created = False
    try:
        tmpl_path = find_template_file(order_dir_base)
        if tmpl_path and tmpl_path.exists():
            _fill_excel_template(tmpl_path, excel_path, folder_name, projects or [], order_notes=order_notes)
            excel_created = True
        else:
            _generate_fallback_excel(excel_path, folder_name, projects or [], order_notes=order_notes)
            excel_created = True
    except Exception as e:
        logger.warning(f"生成 Excel 提测单失败: {e}")

    # Copy Upgrade Explanation Docx file (全部升级说明.docx)
    docx_created = False
    docx_path = None
    try:
        dest_docx = target_dir / "全部升级说明.docx"
        if dest_docx.exists():
            docx_created = False
            docx_path = dest_docx
            logger.info(f"提测目录中已存在全部升级说明.docx，无需再次创建/覆盖: {dest_docx}")
        else:
            docx_tmpl = find_docx_template_file(order_dir_base)
            if docx_tmpl and docx_tmpl.exists():
                import shutil
                shutil.copy2(docx_tmpl, dest_docx)
                docx_created = True
                docx_path = dest_docx
    except Exception as e:
        logger.warning(f"复制升级说明 Word 文档失败: {e}")

    files_msg = f"{excel_name}"
    if docx_created and docx_path:
        files_msg += f", {docx_path.name}"

    return {
        "success": True,
        "dir": str(target_dir),
        "excel": str(excel_path),
        "excel_created": excel_created,
        "docx": str(docx_path) if docx_path else "",
        "docx_created": docx_created,
        "message": f"成功创建目录: {target_dir}（包含文件: {files_msg}）",
    }


def _fill_excel_template(
    tmpl_path: Path,
    output_path: Path,
    folder_name: str,
    projects: List[Dict[str, Any]],
    order_notes:Optional[str] = None,
) -> None:
    """Load the official template file and fill in B6 (folder name) and B12 (changed projects / notes)."""
    import openpyxl

    wb = openpyxl.load_workbook(tmpl_path)

    # Locate worksheet (usually sheet named '测试单' or second worksheet)
    ws = None
    for sheet in wb.worksheets:
        if "测试" in sheet.title or sheet.title != "Export Summary":
            ws = sheet
            break

    if not ws:
        ws = wb.active

    # B6: 产品名称/特殊订单编号
    ws["B6"] = folder_name

    # B12: 产品主要更改项目
    ws["B12"] = _format_change_notes(order_notes, projects or [])

    wb.save(output_path)


def _generate_fallback_excel(
    filepath: Path,
    folder_name: str,
    projects: List[Dict[str, Any]],
    order_notes:Optional[str] = None,
) -> None:
    """Fallback generator if template is missing."""
    try:
        import openpyxl
    except ImportError:
        filepath.touch(exist_ok=True)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试单"

    ws["A2"] = "山东亚华电子股份有限公司"
    ws["A3"] = "测试单"
    ws["A4"] = "编号：YH/D-B04-01-F002                               "
    ws["C4"] = "版本：V1.0     "
    ws["E4"] = "   NO:"
    ws["A5"] = "基本信息"
    ws["A6"] = "产品名称/特殊订单编号"
    ws["B6"] = folder_name
    ws["A7"] = "测试类型/阶段"
    ws["B7"] = "特殊订单"
    ws["A8"] = "产品型号"
    ws["B8"] = "主机：           A10"
    ws["A9"] = "产品版本"
    ws["B9"] = "操作系统:"
    ws["A10"] = "是否需要小批量试制（必选）"
    ws["B10"] = "■不需要            □需要"
    ws["A11"] = "产品主要更改项目"

    ws["A12"] = 1
    ws["B12"] = _format_change_notes(order_notes, projects or [])

    wb.save(filepath)
