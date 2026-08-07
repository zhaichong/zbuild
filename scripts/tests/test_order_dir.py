# -*- coding: utf-8 -*-
import unittest
import tempfile
import shutil
from pathlib import Path
from tools.order_dir import create_order_directory


class TestOrderDirectory(unittest.TestCase):
    def setUp(self):
        self.tmp_dir = Path(tempfile.mkdtemp())

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_creates_folder_and_excel(self):
        order_no = "2026-1396"
        hospital_name = "广州中医药大学顺德医院"
        projects = [{"name": "yarward-web-frontend", "branch": "3.2.1"}]

        res = create_order_directory(self.tmp_dir, order_no, hospital_name, projects)
        self.assertTrue(res["success"])

        expected_folder = self.tmp_dir / "2026-1396-广州中医药大学顺德医院"
        self.assertTrue(expected_folder.exists())
        self.assertTrue(expected_folder.is_dir())

        expected_excel = expected_folder / "2026-1396-广州中医药大学顺德医院医院提测单.xlsx"
        self.assertTrue(expected_excel.exists())

    def test_preserves_existing_docx(self):
        order_no = "2026-1396"
        hospital_name = "广州中医药大学顺德医院"
        expected_folder = self.tmp_dir / "2026-1396-广州中医药大学顺德医院"
        expected_folder.mkdir(parents=True, exist_ok=True)
        docx_file = expected_folder / "全部升级说明.docx"
        docx_file.write_text("existing content", encoding="utf-8")

        res = create_order_directory(self.tmp_dir, order_no, hospital_name, order_notes="测试更新")
        self.assertTrue(res["success"])
        self.assertFalse(res["docx_created"])
        self.assertEqual(docx_file.read_text(encoding="utf-8"), "existing content")

    def test_excel_content_and_template_structure(self):
        import openpyxl
        order_no = "2026-9999"
        hospital_name = "测试医院"
        res = create_order_directory(self.tmp_dir, order_no, hospital_name, order_notes="1、功能修改\n2、样式更新")
        self.assertTrue(res["success"])
        self.assertTrue(res["excel_created"])

        wb = openpyxl.load_workbook(res["excel"])
        ws = None
        for sheet in wb.worksheets:
            if "测试" in sheet.title:
                ws = sheet
                break
        self.assertIsNotNone(ws)
        self.assertEqual(ws["B6"].value, "2026-9999-测试医院")
        self.assertIn("1、功能修改", str(ws["B12"].value))
        self.assertIn("2、样式更新", str(ws["B12"].value))


if __name__ == "__main__":
    unittest.main()
